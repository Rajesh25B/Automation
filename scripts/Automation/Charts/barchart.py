'''
This module creates a new barchart based on the values in a pivot table.
'''

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# load the file
wb = load_workbook('pivot_table.xlsx')

# attach the name of the table
sheet = wb['Report']

# get min & max cols/rows in excel file

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# load the chart
barchart = BarChart()

# calculate data and catagories to be in the BarChart
data = Reference(sheet,
                 min_col=min_column+1,
                 max_col=max_column, min_row=min_row, max_row=max_row)
catagories = Reference(
    sheet,
    min_col=min_column,
    max_col=min_column,
    min_row=min_row+1,
    max_row=max_row
)

# add data and set categories to the barchart
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(catagories)

# place the chart at a location in the excel sheet
sheet.add_chart(barchart, 'B15')

# give a title and style to the chart
barchart.title = 'Sales by Products'
barchart.style = 5

# save the wb
wb.save('barchart.xlsx')
