'''
This module helps writing formulas using python instead of
manual formulas in excel sheets.
'''

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# load the file
wb = load_workbook('pivot_table.xlsx')

# attach the name of the table
sheet = wb['Report']

# get min & max cols/rows in excel file
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# to get SUM of two rows (Electronic Accessories for Male & Female)
# we generally write excel formula as SUM(B7:B8), use python to write this.

# sheet['B8'] = '=SUM(B6:B7)'
# sheet['B8'].style = 'Currency'

# sum for all cells, instead of manually grabbing each cell col & row numbers
# use for loop with min_col, max_col, min_row and max_row
# below cols, rows are based on pivot table.
# get_column_letter(idx) => returns the column name ex: B, C, D...


for i in range(min_column+1, max_column+1):
    # print(i)
    letter = get_column_letter(i)
    sheet[f'{letter}{
        max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style = 'Currency'

wb.save('excel_formulas.xlsx')

# This is how to create formulas in spreadsheet using python.
